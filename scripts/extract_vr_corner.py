#!/usr/bin/env python3
"""Extract SURREALITY 2.0 VR Corner works and prepare web assets."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageFilter


WORKBOOK = Path(
    "/Users/primopan/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_hi0ruq46paz322_6308/msg/file/2026-06/"
    "SURREALITY_2.0_VR Corner(1).xlsx"
)

COS_BASE_URL = "https://surreality-1419044809.cos.ap-hongkong.myqcloud.com/enhanced/"

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

MANUAL_TRANSLATIONS = {
    2: {
        "title_cn": "新纪元的原始人",
        "artist_cn": "Aldo Tambellini / Topos Studio",
        "artist_bio_cn": (
            "Aldo Tambellini（1930-2020）是一位意大利裔美国艺术家，也是电子跨媒介、扩展电影与早期录像艺术的先驱。"
            "他的实践横跨绘画、雕塑、电影、诗歌、表演与电视，并围绕“黑色”作为材料与思想发展出激进的视觉语言。"
            "20 世纪 60 年代，他活跃于纽约实验艺术现场，曾与 Otto Piene、Nam June Paik 等艺术家合作，作品后来在泰特现代美术馆、MoMA、蓬皮杜中心和威尼斯双年展等重要机构展出。"
            "他的创作将艺术、技术与政治紧迫感结合起来，对媒体艺术产生了持久影响。"
        ),
    },
    3: {
        "title_cn": "Astra",
        "artist_cn": "Eliza McNitt",
        "artist_bio_cn": (
            "Eliza McNitt 是一位编剧兼导演，曾入围艾美奖，并获得威尼斯电影节 VR 单元大奖。"
            "她的创作探索科学与艺术在宇宙尺度上的交汇，并长期与宇航员、天体物理学家等科学家合作，讲述人类与宇宙之间的联结。"
            "她的 VR 作品曾亮相 Sundance、SXSW、AFI Fest、Cannes NEXT、Tribeca、Telluride 与 Venice 等电影节。"
        ),
    },
    4: {
        "artist_cn": "Elke Reinhuber & Benjamin Seide",
    },
    5: {
        "title_cn": "作为记忆是什么样的体验？",
    },
    6: {
        "title_cn": "我软件中的虫",
        "artist_cn": "Alina Nazmeeva",
        "artist_bio_cn": (
            "Alina Nazmeeva 是出生于鞑靼斯坦的媒体艺术家，以模拟作为媒介与批判场域。"
            "她现任 UIUC 计算方向助理教授，作品曾在威尼斯建筑双年展、洛杉矶建筑与设计博物馆、Beall Center 和 Slamdance 等展出。"
            "她曾在 Art Omi、密歇根大学、加拿大建筑中心和 Strelka Institute 驻留，并曾在 Harvard GSD、Duke University、Parsons、ISEA、SIGGRAPH 等机构与会议发表作品。"
        ),
    },
    7: {
        "title_cn": "Respire 2.0 沉浸式互动 VR 作品",
        "artist_cn": (
            "艺术总监：Philippe Pasquier、Mirjana Prpa、Katerina Semenko；"
            "3D 艺术家：Katerina Semenko；交互设计：Mirjana Prpa；程序：Szeka Tse；"
            "研究助理：Dhruv Adhia；声音：Philippe Pasquier、Keon Leigh、Ge Liu"
        ),
        "artist_bio_cn": (
            "Philippe Pasquier 是 SFU 教授及 Metacreation Lab 主任，作为科学家兼艺术家发表 200 余篇论文，作品曾在蓬皮杜中心和 Ars Electronica 展出。"
            "Mirjana Prpa 博士是香港科技大学（广州）助理教授，研究 AI、人机交互与以人为中心的 XR 设计，并曾获 ACM CHI 奖项。"
            "Katerina Semenko 是独立 phygital 艺术家与设计师，曾获 IDA Design of the Year、A'Design Award 等 30 余项荣誉。"
            "团队还包括提供技术支持的程序员 Szeka Tse、研究 AI 驱动打击乐的 SFU 博士候选人 Keon Ju Maverick Lee、研究生成音乐的 SFU 研究者与游戏作曲家 Ge Liu，以及拥有 15 年以上 XR 与空间计算经验的创意技术专家 Dhruv Adhia。"
        ),
    },
    10: {
        "artist_cn": "Benedict Yu",
    },
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if text in {"#VALUE!", "图片"}:
        return ""
    return text


def clean_year(value: Any) -> str:
    text = clean(value)
    match = re.search(r"(19|20)\d{2}", text.replace(",", ""))
    return match.group(0) if match else ""


def slugify(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "vr-work"


def sort_key(title: str) -> str:
    return title.strip().strip('"“”').lower()


def load_image_cell_map(workbook_path: Path) -> dict[int, str]:
    with zipfile.ZipFile(workbook_path) as zf:
        rel_root = ET.fromstring(zf.read("xl/richData/_rels/richValueRel.xml.rels"))
        rels = {
            rel.attrib["Id"]: rel.attrib["Target"].replace("../", "xl/")
            for rel in rel_root
        }

        rv_root = ET.fromstring(zf.read("xl/richData/rdrichvalue.xml"))
        rv_to_media: dict[int, str] = {}
        for index, rv in enumerate(rv_root):
            first_value = int(list(rv)[0].text or "0")
            rv_to_media[index] = rels[f"rId{first_value + 1}"]

        vm_to_media = {str(index + 1): rv_to_media[index] for index in rv_to_media}
        root = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))
        result: dict[int, str] = {}
        for cell in root.findall(".//main:c", NS):
            ref = cell.attrib.get("r", "")
            vm = cell.attrib.get("vm")
            if ref.startswith("O") and vm:
                result[int(ref[1:])] = vm_to_media[vm]
        return result


def extract_media(workbook_path: Path, media_name: str, destination: Path) -> None:
    with zipfile.ZipFile(workbook_path) as zf:
        destination.write_bytes(zf.read(media_name))


def needs_ai_upscale(width: int, height: int, size_bytes: int) -> bool:
    pixels = width * height
    bytes_per_pixel = size_bytes / max(pixels, 1)
    return min(width, height) < 1100 or bytes_per_pixel < 0.16


def run_super_resolution(source: Path, target: Path, model_path: Path | None) -> bool:
    if not model_path or not model_path.exists():
        return False
    try:
        import cv2  # type: ignore
    except Exception:
        return False

    image = cv2.imread(str(source), cv2.IMREAD_COLOR)
    if image is None:
        return False

    sr = cv2.dnn_superres.DnnSuperResImpl_create()
    sr.readModel(str(model_path))
    sr.setModel("fsrcnn", 2)
    upscaled = sr.upsample(image)
    return bool(cv2.imwrite(str(target), upscaled))


def enhance_image(source: Path, target: Path, model_path: Path | None) -> dict[str, Any]:
    with Image.open(source) as original:
        original.load()
        width, height = original.size
        source_size = source.stat().st_size
        do_ai = needs_ai_upscale(width, height, source_size)

    temp_png = target.with_suffix(".ai-tmp.png")
    used_ai = do_ai and run_super_resolution(source, temp_png, model_path)
    working_source = temp_png if used_ai else source

    with Image.open(working_source) as image:
        image.load()
        if image.mode not in {"RGB", "RGBA"}:
            image = image.convert("RGBA" if "A" in image.getbands() else "RGB")

        long_edge = max(image.size)
        if long_edge > 3200:
            scale = 3200 / long_edge
            next_size = (round(image.size[0] * scale), round(image.size[1] * scale))
            image = image.resize(next_size, Image.Resampling.LANCZOS)

        image = image.filter(ImageFilter.UnsharpMask(radius=1.1, percent=115, threshold=4))
        image = ImageEnhance.Contrast(image).enhance(1.035)
        image.save(target, "WEBP", quality=92, method=6)
        output_dimensions = Image.open(target).size

    if temp_png.exists():
        temp_png.unlink()

    return {
        "used_ai_super_resolution": used_ai,
        "source_width": width,
        "source_height": height,
        "source_bytes": source_size,
        "output_width": output_dimensions[0],
        "output_height": output_dimensions[1],
        "output_bytes": target.stat().st_size,
    }


def build_records(workbook_path: Path, image_map: dict[int, str]) -> list[dict[str, Any]]:
    worksheet = load_workbook(workbook_path, data_only=False)["VR"]
    records: list[dict[str, Any]] = []

    for row in range(2, worksheet.max_row + 1):
        title_en = clean(worksheet.cell(row, 8).value)
        artist_en = clean(worksheet.cell(row, 1).value)
        if not title_en and not artist_en:
            continue

        translation = MANUAL_TRANSLATIONS.get(row, {})
        title_cn = clean(worksheet.cell(row, 9).value) or translation.get("title_cn", "")
        artist_cn = clean(worksheet.cell(row, 2).value) or translation.get("artist_cn", "")
        description_en = clean(worksheet.cell(row, 12).value) or clean(worksheet.cell(row, 10).value)
        description_cn = clean(worksheet.cell(row, 13).value) or clean(worksheet.cell(row, 11).value)
        artist_bio_en = clean(worksheet.cell(row, 5).value) or clean(worksheet.cell(row, 3).value)
        artist_bio_cn = clean(worksheet.cell(row, 6).value) or clean(worksheet.cell(row, 4).value) or translation.get("artist_bio_cn", "")

        records.append(
            {
                "source_sheet": "VR",
                "source_row": row,
                "title_en": title_en,
                "title_cn": title_cn or title_en,
                "artist_en": artist_en,
                "artist_cn": artist_cn or artist_en,
                "description_en": description_en,
                "description_cn": description_cn or description_en,
                "artist_bio_en": artist_bio_en,
                "artist_bio_cn": artist_bio_cn or artist_bio_en,
                "year": clean_year(worksheet.cell(row, 14).value),
                "source_media": image_map[row],
                "sort_key": sort_key(title_en or title_cn),
            }
        )

    records.sort(key=lambda item: item["sort_key"])
    for index, record in enumerate(records, 1):
        slug = slugify(record["title_en"])
        image_file = f"vr-{index:02d}-{slug}.webp"
        record["id"] = index
        record["image_file"] = image_file
        record["poster_url"] = COS_BASE_URL + image_file
        record["local_poster_url"] = "/vr-corner-2026/enhanced/" + image_file

    return records


def write_manifest(records: list[dict[str, Any]], manifest_path: Path) -> None:
    manifest_path.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    fields = [
        "id",
        "sort_key",
        "source_sheet",
        "source_row",
        "title_en",
        "title_cn",
        "artist_en",
        "artist_cn",
        "source_media",
        "image_file",
        "poster_url",
        "used_ai_super_resolution",
        "source_width",
        "source_height",
        "output_width",
        "output_height",
    ]
    with manifest_path.with_suffix(".csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fields)
        writer.writeheader()
        for record in records:
            writer.writerow({field: record.get(field, "") for field in fields})


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    parser.add_argument("--output-root", type=Path, default=Path("exports/vr-corner-2026"))
    parser.add_argument("--public-image-root", type=Path, default=Path("public/vr-corner-2026/enhanced"))
    parser.add_argument("--public-data", type=Path, default=Path("public/data/vr-corner-2026.json"))
    parser.add_argument("--model", type=Path, default=Path("/tmp/mr-models/FSRCNN_x2.pb"))
    args = parser.parse_args()

    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)

    image_map = load_image_cell_map(args.workbook)
    records = build_records(args.workbook, image_map)

    originals = args.output_root / "originals"
    enhanced = args.output_root / "enhanced"
    originals.mkdir(parents=True, exist_ok=True)
    enhanced.mkdir(parents=True, exist_ok=True)
    args.public_image_root.mkdir(parents=True, exist_ok=True)
    args.public_data.parent.mkdir(parents=True, exist_ok=True)

    for record in records:
        original_suffix = Path(record["source_media"]).suffix.lower()
        original_path = originals / record["image_file"].replace(".webp", original_suffix)
        enhanced_path = enhanced / record["image_file"]
        public_path = args.public_image_root / record["image_file"]

        extract_media(args.workbook, record["source_media"], original_path)
        metrics = enhance_image(original_path, enhanced_path, args.model)
        shutil.copy2(enhanced_path, public_path)
        record.update(metrics)

    public_records = [
        {key: value for key, value in record.items() if key not in {"source_media", "source_bytes", "output_bytes"}}
        for record in records
    ]
    args.public_data.write_text(json.dumps(public_records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_manifest(records, args.output_root / "manifest.json")

    print(json.dumps(
        {
            "records": len(records),
            "unique_images": len({record["image_file"] for record in records}),
            "used_ai_super_resolution": sum(bool(record.get("used_ai_super_resolution")) for record in records),
            "order": [record["title_en"] for record in records],
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()

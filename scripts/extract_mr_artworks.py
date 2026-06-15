#!/usr/bin/env python3
"""Extract SURREALITY 2.0 MR artworks and prepare web image assets."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageFilter


WORKBOOK = Path(
    "/Users/primopan/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_hi0ruq46paz322_6308/msg/file/2026-06/"
    "SURREALITY_2.0_两个校园MR.xlsx"
)

COS_BASE_URL = "https://surreality-1419044809.cos.ap-hongkong.myqcloud.com/mr-artworks-2026/enhanced/"

SHEET_META = {
    "GZ": {"campus": "guangzhou", "campus_label_en": "Guangzhou", "campus_label_cn": "广州", "prefix": "gz"},
    "CWB": {"campus": "hong-kong", "campus_label_en": "Hong Kong", "campus_label_cn": "香港", "prefix": "hk"},
}

AREA_LABELS = {
    "展区一": 1,
    "展区二": 2,
    "展区三": 3,
    "展区四": 4,
}

MANUAL_TRANSLATIONS = {
    ("GZ", 3): {"title_cn": "Abject Space", "artist_cn": "Adam Nash"},
    ("GZ", 4): {"title_cn": "共振之弧", "artist_cn": "UnCalculated Studio"},
    ("GZ", 9): {"title_cn": "季节之云", "artist_cn": "Arnaud Laffond"},
    ("GZ", 10): {"title_cn": "我软件中的虫", "artist_cn": "Alina Nazmeeva"},
    ("GZ", 13): {"title_cn": "GFP Bunny", "artist_cn": "Eduardo Kac"},
    ("GZ", 16): {"title_cn": "算法 AI 数字艺术 2026", "artist_cn": "Kadine James / The Immersive KIND"},
    ("CWB", 6): {"title_cn": "拉斯·阿维查斯", "artist_cn": "Violeta Ayala"},
    ("CWB", 7): {"title_cn": "GFP Bunny", "artist_cn": "Eduardo Kac"},
    ("CWB", 11): {"title_cn": "双重本性", "artist_cn": "Katerina Semenko"},
}

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if text in {"#VALUE!", "图片"}:
        return ""
    return text


def clean_year(value: Any) -> str:
    text = clean(value)
    match = re.search(r"(19|20)\d{2}", text.replace(",", ""))
    return match.group(0) if match else ""


def slugify(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "artwork"


def load_image_cell_map(workbook_path: Path) -> dict[tuple[str, int], str]:
    with zipfile.ZipFile(workbook_path) as zf:
        rel_root = ET.fromstring(zf.read("xl/richData/_rels/richValueRel.xml.rels"))
        rels = {
            rel.attrib["Id"]: rel.attrib["Target"].replace("../", "xl/")
            for rel in rel_root
        }

        rv_root = ET.fromstring(zf.read("xl/richData/rdrichvalue.xml"))
        rv_to_media: dict[int, str] = {}
        for index, rv in enumerate(rv_root):
            first_value = int(list(rv)[0].text or "0")
            rv_to_media[index] = rels[f"rId{first_value + 1}"]

        vm_to_media = {str(index + 1): rv_to_media[index] for index in rv_to_media}

        result: dict[tuple[str, int], str] = {}
        for sheet_xml, sheet_name in [
            ("xl/worksheets/sheet1.xml", "GZ"),
            ("xl/worksheets/sheet2.xml", "CWB"),
        ]:
            root = ET.fromstring(zf.read(sheet_xml))
            for cell in root.findall(".//main:c", NS):
                ref = cell.attrib.get("r", "")
                vm = cell.attrib.get("vm")
                if ref.startswith("P") and vm:
                    result[(sheet_name, int(ref[1:]))] = vm_to_media[vm]
        return result


def extract_media(workbook_path: Path, media_name: str, destination: Path) -> None:
    with zipfile.ZipFile(workbook_path) as zf:
        destination.write_bytes(zf.read(media_name))


def needs_ai_upscale(width: int, height: int, size_bytes: int) -> bool:
    pixels = width * height
    bytes_per_pixel = size_bytes / max(pixels, 1)
    return min(width, height) < 1100 or bytes_per_pixel < 0.16


def run_super_resolution(source: Path, target: Path, model_path: Path | None) -> bool:
    if not model_path or not model_path.exists():
        return False

    try:
        import cv2  # type: ignore
    except Exception:
        return False

    image = cv2.imread(str(source), cv2.IMREAD_COLOR)
    if image is None:
        return False

    sr = cv2.dnn_superres.DnnSuperResImpl_create()
    sr.readModel(str(model_path))
    sr.setModel("fsrcnn", 2)
    upscaled = sr.upsample(image)
    ok = cv2.imwrite(str(target), upscaled)
    return bool(ok)


def enhance_image(source: Path, target: Path, model_path: Path | None) -> dict[str, Any]:
    with Image.open(source) as original:
        original.load()
        width, height = original.size
        source_size = source.stat().st_size
        do_ai = needs_ai_upscale(width, height, source_size)

    temp_png = target.with_suffix(".ai-tmp.png")
    used_ai = False
    if do_ai:
        used_ai = run_super_resolution(source, temp_png, model_path)

    working_source = temp_png if used_ai else source
    with Image.open(working_source) as image:
        image.load()
        if image.mode not in {"RGB", "RGBA"}:
            image = image.convert("RGBA" if "A" in image.getbands() else "RGB")

        long_edge = max(image.size)
        if long_edge > 3200:
            scale = 3200 / long_edge
            next_size = (round(image.size[0] * scale), round(image.size[1] * scale))
            image = image.resize(next_size, Image.Resampling.LANCZOS)

        image = image.filter(ImageFilter.UnsharpMask(radius=1.1, percent=115, threshold=4))
        image = ImageEnhance.Contrast(image).enhance(1.035)

        if image.mode == "RGBA":
            image.save(target, "WEBP", quality=92, method=6, lossless=False)
        else:
            image.save(target, "WEBP", quality=92, method=6)

        output_size = target.stat().st_size
        output_dimensions = Image.open(target).size

    if temp_png.exists():
        temp_png.unlink()

    return {
        "used_ai_super_resolution": used_ai,
        "source_width": width,
        "source_height": height,
        "source_bytes": source_size,
        "output_width": output_dimensions[0],
        "output_height": output_dimensions[1],
        "output_bytes": output_size,
    }


def build_records(workbook_path: Path, image_map: dict[tuple[str, int], str]) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=False)
    records: list[dict[str, Any]] = []
    area_counts: dict[tuple[str, int], int] = {}

    for worksheet in workbook.worksheets:
        if worksheet.title not in SHEET_META:
            continue
        meta = SHEET_META[worksheet.title]
        for row in range(2, worksheet.max_row + 1):
            area_text = clean(worksheet.cell(row, 1).value)
            if area_text not in AREA_LABELS:
                continue

            area = AREA_LABELS[area_text]
            area_key = (meta["campus"], area)
            area_counts[area_key] = area_counts.get(area_key, 0) + 1
            order = area_counts[area_key]

            title_en = clean(worksheet.cell(row, 9).value)
            title_cn = clean(worksheet.cell(row, 10).value)
            artist_en = clean(worksheet.cell(row, 2).value)
            artist_cn = clean(worksheet.cell(row, 3).value)
            description_en = clean(worksheet.cell(row, 13).value) or clean(worksheet.cell(row, 11).value)
            description_cn = clean(worksheet.cell(row, 14).value) or clean(worksheet.cell(row, 12).value)
            artist_bio_en = clean(worksheet.cell(row, 6).value) or clean(worksheet.cell(row, 4).value)
            artist_bio_cn = clean(worksheet.cell(row, 7).value) or clean(worksheet.cell(row, 5).value)

            translation = MANUAL_TRANSLATIONS.get((worksheet.title, row), {})
            title_cn = title_cn or translation.get("title_cn", "")
            artist_cn = artist_cn or translation.get("artist_cn", "")

            year = clean_year(worksheet.cell(row, 15).value)
            media_name = image_map[(worksheet.title, row)]
            slug = slugify(title_en or title_cn)
            image_file = f"{meta['prefix']}-area{area}-{order:02d}-{slug}.webp"

            records.append(
                {
                    "id": f"{meta['prefix']}-area{area}-{order:02d}",
                    "campus": meta["campus"],
                    "campus_label_en": meta["campus_label_en"],
                    "campus_label_cn": meta["campus_label_cn"],
                    "area": area,
                    "order": order,
                    "source_sheet": worksheet.title,
                    "source_row": row,
                    "title_en": title_en,
                    "title_cn": title_cn,
                    "artist_en": artist_en,
                    "artist_cn": artist_cn,
                    "description_en": description_en,
                    "description_cn": description_cn,
                    "artist_bio_en": artist_bio_en,
                    "artist_bio_cn": artist_bio_cn,
                    "year": year,
                    "source_media": media_name,
                    "image_file": image_file,
                    "image_url": COS_BASE_URL + image_file,
                    "local_image_url": "/mr-artworks-2026/enhanced/" + image_file,
                }
            )
    return records


def write_manifest(records: list[dict[str, Any]], manifest_path: Path) -> None:
    manifest_path.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    csv_path = manifest_path.with_suffix(".csv")
    fields = [
        "id",
        "campus",
        "area",
        "order",
        "source_sheet",
        "source_row",
        "title_en",
        "title_cn",
        "artist_en",
        "artist_cn",
        "source_media",
        "image_file",
        "image_url",
        "used_ai_super_resolution",
        "source_width",
        "source_height",
        "output_width",
        "output_height",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fields)
        writer.writeheader()
        for record in records:
            writer.writerow({field: record.get(field, "") for field in fields})


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    parser.add_argument("--output-root", type=Path, default=Path("exports/mr-artworks-2026"))
    parser.add_argument("--public-image-root", type=Path, default=Path("public/mr-artworks-2026/enhanced"))
    parser.add_argument("--public-data", type=Path, default=Path("public/data/mr-artworks-2026.json"))
    parser.add_argument("--model", type=Path, default=Path("/tmp/mr-models/FSRCNN_x2.pb"))
    args = parser.parse_args()

    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)

    image_map = load_image_cell_map(args.workbook)
    records = build_records(args.workbook, image_map)

    originals = args.output_root / "originals"
    enhanced = args.output_root / "enhanced"
    originals.mkdir(parents=True, exist_ok=True)
    enhanced.mkdir(parents=True, exist_ok=True)
    args.public_image_root.mkdir(parents=True, exist_ok=True)
    args.public_data.parent.mkdir(parents=True, exist_ok=True)

    media_to_file: dict[str, str] = {}
    for record in records:
        source_media = record["source_media"]
        image_file = record["image_file"]
        if source_media in media_to_file:
            record["image_file"] = media_to_file[source_media]
            record["image_url"] = COS_BASE_URL + record["image_file"]
            record["local_image_url"] = "/mr-artworks-2026/enhanced/" + record["image_file"]
            continue
        media_to_file[source_media] = image_file

        original_suffix = Path(source_media).suffix.lower()
        original_path = originals / image_file.replace(".webp", original_suffix)
        enhanced_path = enhanced / image_file
        public_path = args.public_image_root / image_file

        extract_media(args.workbook, source_media, original_path)
        metrics = enhance_image(original_path, enhanced_path, args.model)
        shutil.copy2(enhanced_path, public_path)
        record.update(metrics)

    for record in records:
        if "used_ai_super_resolution" not in record:
            first = next(item for item in records if item["source_media"] == record["source_media"] and "used_ai_super_resolution" in item)
            for key in [
                "used_ai_super_resolution",
                "source_width",
                "source_height",
                "source_bytes",
                "output_width",
                "output_height",
                "output_bytes",
            ]:
                record[key] = first[key]

    public_records = [
        {key: value for key, value in record.items() if key not in {"source_media", "source_bytes", "output_bytes"}}
        for record in records
    ]
    args.public_data.write_text(json.dumps(public_records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_manifest(records, args.output_root / "manifest.json")

    counts: dict[str, int] = {}
    for record in records:
        key = f"{record['campus']}-area{record['area']}"
        counts[key] = counts.get(key, 0) + 1

    print(json.dumps({"records": len(records), "unique_images": len(media_to_file), "counts": counts}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
